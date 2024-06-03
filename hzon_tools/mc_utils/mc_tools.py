import hashlib
import io
import os
import uuid

import openpyxl
import datetime
import datetime
import socket
import requests
import time

from datetime import datetime as dt

import requests

"""
connect = pymysql.connect(
    host=glv['mysql_config'].get('host'),
    port=3306,
    db='operating-management',
    user=glv['mysql_config'].get('user'),
    passwd=glv['mysql_config'].get('passwd'),
    charset='utf8',
    use_unicode=True,
    cursorclass=pymysql.cursors.DictCursor  # 设置 cursorclass 为 DictCursor
)

# redis连接池
redis_client = redis.StrictRedis(host='ip', password="passwd", port=6379, db=3)

import platform

isLinux = False if platform.system() == "Windows" else True

def save_data(self,lis):
    print(lis)

    try:
        self.connect.ping(reconnect=True)
        cursor = self.connect.cursor()
        sql = '''
                    insert ignore into table_name (
        字段1,字段2,字段3,字段4,字段5,字段6,字段7,字段8,字段9,字段10
        ) values (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                  %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,
                  %s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            '''
        cursor.executemany(sql, lis)
        self.connect.commit()
        cursor.close()
        print("插入成功")
    except Exception as e:
        print(e)
        self.connect.rollback()
        print("插入失败")
"""

def get_yesterday():
    """
    获取昨天的日期
    :return:
    """
    today = dt.today().date()  # 今天
    yestoday = dt.strftime(today + datetime.timedelta(days=-1), '%Y-%m-%d')
    return yestoday


def parse_excel(file, dir):
    """
    读取excel文件
    :param file:
    :param dir:
    :return:
    """
    get_date = get_yesterday()
    print("开始处理文件: %s" % file)
    files = os.path.join(dir, file)
    if not os.path.exists(files):
        print("当前文件不存在%s" % files)

    try:
        with open(files, "rb") as f:  # rb模式：以字节(二进制)方式读取文件中的数据
            in_mem_file = io.BytesIO(f.read())
        book = openpyxl.load_workbook(in_mem_file)
        sheet = book.active
        max_row = sheet.max_row
        max_column = sheet.max_column
        print("max_row: %d, max_column: %d" % (max_row, max_column))
        result = []
        heads = []
        for column in range(max_column):
            # 读取的话行列是从（1，1）开始
            heads.append(sheet.cell(1, column + 1).value)
        for row in range(max_row):
            if row == 0:
                continue
            one_line = {}
            for column in range(max_column):
                # print("row: %d, column: %d" % (row, column))
                # 读取第二行开始每一个数据
                k = heads[column]
                cell = sheet.cell(row + 1, column + 1)
                value = cell.value if cell.value != "-" else None
                one_line[k] = value
            result.append(one_line)

        book.close()
        return result


    except Exception as e:
        print("处理文件: %s 失败, 错误信息: %s" % (file, e))


def run_spider_task():
    """
    使用api运行crawlab中的爬虫程序
    """
    # 爬虫程序开始之前,将招聘爬虫状态改成1

    # redis_client.set("Resume:active", 1)
    headers = {
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "zh-CN,zh;q=0.9",
        "Authorization": "${token}",
        "Connection": "keep-alive",
        "Content-Type": "application/json",
        "Origin": "http://127.0.0.1:8080",
        "Referer": "http://127.0.0.1:8080/",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    }
    url = "http://127.0.0.1:8080/api/spiders/658d4a97cd698bc0ab7d5770/run"
    data = {
        "mode": "random",
        "cmd": "python 招聘项目/v2Online/resume_feapder_spider.py",
        "param": "",
        "priority": 5
    }
    response = requests.post(url, headers=headers, json=data, verify=False)
    print(response.text)
    print(response)


def get_date(years=[2023, 2024]):
    """
    循环获取years列表中的年份,然后获取年份所有的日期,用列表的形式返回所有日期
    """
    date_list = []
    # 获取今天的日期
    today = datetime.datetime.now().strftime("%Y-%m-%d")
    # 遍历years列表
    for year in years:
        start_date = datetime.date(year, 1, 1)
        end_date = datetime.date(year, 12, 31)
        delta = datetime.timedelta(days=1)

        current_date = start_date
        while current_date <= end_date:
            # print(current_date.strftime('%Y-%m-%d'))
            date_list.append(current_date.strftime('%Y-%m-%d'))
            if current_date.strftime('%Y-%m-%d') == today:
                return date_list[:-1]
            current_date += delta


def get_need_spider_date(_date_list):
    """
    获取需要爬取的日期列表
    :param _date_list: 已经爬取过的日期列表
    :return:
    """
    # 跑全量数据
    # return ['2023-08-31', '2023-08-30', '2023-08-29',...]
    # return sorted(get_date(), reverse=True)

    # with connect.cursor() as cursor:
    #     sql = """
    #         select data_date from order_information group by data_date order by data_date desc
    #         """
    #     cursor.execute(sql)
    #     result = cursor.fetchall()
    #     _date_list = [i["data_date"].strftime("%Y-%m-%d") for i in result]
    # 求date_list和result的差集
    r = list(set(get_date()).difference(set(_date_list)))
    return sorted(r, reverse=False)


def get_cur_start_end_date(self):
    """
    获取今天的开始时间戳和结束时间戳,长度要求为13位
    :return:
    """
    today = datetime.date.today()
    today_start = int(time.mktime(time.strptime(str(today - datetime.timedelta(days=1)), '%Y-%m-%d'))) * 1000
    today_end = today_start + 86400000 - 1
    return str(today_start), str(today_end)

def timestamp_to_str(timestamp):
    """
    将时间戳转换成字符串
    :param timestamp:
    :return:
    """
    total_time = time.time() - timestamp
    # 计算两个时间段的时间差,精确到时分秒
    m, s = divmod(total_time, 60)
    h, m = divmod(m, 60)
    d, h = divmod(h, 24)
    timeArray = time.localtime(timestamp)
    str_date = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)
    return "运行开始时间:{}".format(str_date), "运行总时长:{}天{}小时{}分{}秒".format(int(d), int(h), int(m),
                                                                                      int(s))


def send_wx_message_by_txt(args):
    url = "http://127.0.0.1:8101/microser/send_wx_message"
    payload = {
        "bot_url": "https://qyapi.weixin.qq.com/cgi-bin/webhook/send?key=${key}",
        "content": "项目名称:{}\n错误信息:{}".format(args["project_name"], args["exception"])
    }
    headers = {
        "kw": "",
        "content-type": "application/json"
    }

    response = requests.request("POST", url, json=payload, headers=headers)

    print(response.text)


def send_wx_message_by_md(args):
    """
    以markdown的形式发送企微消息
    :param args:{
        "bot_url":"",
        "user_ids":"",
        "project_name":"",
        "msg":""
        }
    :return:
    """

    def get_ip_address():
        try:
            # 使用socket库获取本机主机名
            host_name = socket.gethostname()
            # 获取主机名对应的IP地址
            ip_address = socket.gethostbyname(host_name)
            return ip_address
        except socket.error as e:
            print(f"获取IP地址时出错：{e}")
            return None

    # 获取本机计算机名
    def get_computer_name():
        try:
            # 使用socket库获取本机主机名
            host_name = socket.gethostname()
            return host_name
        except socket.error as e:
            print(f"获取计算机名时出错：{e}")
            return None

    def send_msg(args):
        url = args.get("bot_url")
        user_ids = args.get('user_ids')
        user_ids_str = "".join([f"<@{user_id}>" for user_id in user_ids]) if user_ids else ""
        payload = {
            "msgtype": "markdown",
            # "content": "项目名称: {}\n本机IP地址: {}\n本机计算机名: {}\n告警信息: {}\n时间:{}".format(
            # 	args.get("project_name"), args.get("ip_address"), args.get("computer_name"), args.get("msg"),
            # 	args.get("now_time"))
            "markdown": {
                "content": "**`项目名称`**: {}\n**`本机IP地址`**: {}\n**`本机计算机名`**: {}\n**`告警信息`**: {}\n**`时间`**: {}\n{}".format(
                    args.get("project_name"), args.get("ip_address"), args.get("computer_name"), args.get("msg"),
                    args.get("now_time"), user_ids_str)

            }
        }
        headers = {
            'Content-Type': 'application/json'
        }

        response = requests.request("POST", url, headers=headers, json=payload)

        print(response.text)

    args["ip_address"] = get_ip_address()
    args["computer_name"] = get_computer_name()
    args["now_time"] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    send_msg(args)


def get_random_file_name():
    """
    随机获取一个文件名
    """
    uid = uuid.uuid4()
    return str(uid)

def md5_encode(self, string):
    """
    md5加密
    :param string:
    :return:
    """
    md5 = hashlib.md5()
    md5.update(string.encode('utf-8'))
    return md5.hexdigest()